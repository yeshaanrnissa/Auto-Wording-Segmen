import io
import re
import openpyxl
import pandas as pd
import streamlit as st
import math
from docx import Document


st.set_page_config(page_title="Auto Wording Segmen", layout="wide")

SEGMEN_LIST = ["Prioritas", "Payroll", "Micro", "Medium Entrepreneur"]

DEFAULT_KOLOM = {
    "Jumlah NoA": "E",
    "MTD NoA": "F",
    "% Penetrasi": "G",
    "Penjualan": "H",
    "Rata-rata Ticket Size": "J",
    "Δ MTD NoA": "N",
    "Δ MTD Penjualan": "P",
    "Δ MTD Ticket Size": "R",
}

def format_ribuan(angka):
    if angka is None or angka == "":
        return "0"
    try:
        # Pisahkan tanda minus jika ada
        string_angka = str(float(angka))
        negatif = "-" if string_angka.startswith("-") else ""
        string_angka = string_angka.lstrip("-")
        
        # Ambil bagian bulatnya saja sebelum koma desimal
        bagian_bulat = string_angka.split('.')[0]
        
        # Format ribuan untuk bagian bulat
        ribuan = f"{int(bagian_bulat):,}".replace(",", ".")
        return f"{negatif}{ribuan}"
    except (ValueError, TypeError):
        return "0"


def format_desimal(angka, digit=2):
    if angka is None or angka == "":
        return "0"
    try:
        # Menggunakan format bawaan Python (:,.2f) yang otomatis membulatkan seperti Excel
        # Ditambah trik replace standar Indonesia (Titik jadi X, Koma jadi Titik, X jadi Koma)
        return f"{float(angka):,.{digit}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "0"


def format_percentage(angka, digit=2):
    if angka is None or angka == "":
        return "0%"
    try:
        # Mengubah ke float dan dikalikan 100 untuk persentase
        nilai = float(angka) * 100
        
        # Menggunakan format bawaan (:,.f) agar otomatis membulatkan seperti Excel
        string_format = f"{nilai:,.{digit}f}"
        
        # Ubah format ribuan dan desimal sesuai standar Indonesia
        return string_format.replace(",", "X").replace(".", ",").replace("X", ".") + "%"
    except (ValueError, TypeError):
        return "0%"

BULAN_INDONESIA = [
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember"
]


def format_tanggal_indonesia(tanggal):
    return f"{tanggal.day} {BULAN_INDONESIA[tanggal.month - 1]} {tanggal.year}"



def format_tanggal_indonesia(tanggal):
    return f"{tanggal.day} {BULAN_INDONESIA[tanggal.month - 1]} {tanggal.year}"


# Search nama segmen

def cari_baris_segmen(sheet, keyword_mapping, max_row_scan=1000, max_col_scan=15):
    hasil = {s: None for s in SEGMEN_LIST}
    max_row = min(sheet.max_row, max_row_scan)
    max_col = min(sheet.max_column, max_col_scan)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = sheet.cell(row=row, column=col).value
            if not isinstance(val, str):
                continue
            val_clean = val.strip().lower()
            for segmen, keyword in keyword_mapping.items():
                if hasil[segmen] is None and val_clean == keyword.strip().lower():
                    hasil[segmen] = row
    return hasil


def cari_baris_grand_total(sheet, label_rows, grand_total_keyword="Grand Total",
                            max_row_scan=1000, max_col_scan=15):
    hasil = {s: None for s in label_rows}
    label_terurut = sorted(
        [(s, r) for s, r in label_rows.items() if r], key=lambda x: x[1]
    )
    max_row = min(sheet.max_row, max_row_scan)
    max_col = min(sheet.max_column, max_col_scan)
    kw = grand_total_keyword.strip().lower()

    for idx, (segmen, label_row) in enumerate(label_terurut):
        batas_bawah = (
            label_terurut[idx + 1][1] - 1 if idx + 1 < len(label_terurut) else max_row
        )
        for row in range(label_row + 1, batas_bawah + 1):
            for col in range(1, max_col + 1):
                val = sheet.cell(row=row, column=col).value
                if isinstance(val, str) and kw in val.strip().lower():
                    hasil[segmen] = row
                    break
            if hasil[segmen]:
                break
    return hasil


def get_row_data(sheet, row, kolom_mapping):
    if row is None or row < 1:
        return {k: None for k in kolom_mapping}
    return {
        label: sheet[f"{kolom}{row}"].value
        for label, kolom in kolom_mapping.items()
    }

# Replace paragraf di Word
def set_paragraph_text(para, new_text):
    if not para.runs:
        return
    para.runs[0].text = new_text
    for run in para.runs[1:]:
        run.text = ""

def proses_dokumen(doc, data_per_segmen, tanggal_laporan=None):
    current_section = None
    jumlah_diubah = 0
    log = []

    # =========================
    # FORMAT TANGGAL
    # =========================
    tanggal_text = None

    if tanggal_laporan:
        tanggal_text = format_tanggal_indonesia(tanggal_laporan)

    # Regex untuk mencari:
    # sd 7 Agustus 2026
    # sd 10 Agustus 2026
    # sd 1 Januari 2027
    pola_tanggal = re.compile(
        r"sd\s+\d{1,2}\s+"
        r"(?:Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|"
        r"September|Oktober|November|Desember)"
        r"\s+\d{4}",
        flags=re.IGNORECASE
    )

    for para in doc.paragraphs:

        # Simpan teks awal
        text_sebelum = para.text.strip()

        # =========================
        # AUTOMASI TANGGAL
        # =========================
        if tanggal_text and text_sebelum:

            new_text_tanggal = pola_tanggal.sub(
                f"sd {tanggal_text}",
                para.text
            )

            if new_text_tanggal != para.text:
                log.append(
                    f"Tanggal: {text_sebelum!r} -> {new_text_tanggal.strip()!r}"
                )

                set_paragraph_text(para, new_text_tanggal)

            # Update text setelah tanggal diganti
            text = new_text_tanggal.strip()

        else:
            text = text_sebelum

        # =========================
        # DETEKSI SEGMEN
        # =========================
        if text == "Prioritas":
            current_section = data_per_segmen.get("Prioritas")
            continue

        elif text == "Payroll":
            current_section = data_per_segmen.get("Payroll")
            continue

        elif text == "Micro":
            current_section = data_per_segmen.get("Micro")
            continue

        elif text == "Medium Entrepreneur":
            current_section = data_per_segmen.get("Medium Entrepreneur")
            continue

        if current_section is None or not para.runs:
            continue

        # =========================
        # AUTOMASI WORDING
        # =========================
        new_text = None

        if text.startswith("Jumlah NoA"):
            new_text = (
                f"Jumlah NoA: "
                f"{format_ribuan(current_section['Jumlah NoA'])} NoA"
            )

        elif text.startswith("MTD NoA"):
            new_text = (
                f"MTD NoA: "
                f"{format_ribuan(current_section['MTD NoA'])} NoA"
            )

        elif text.startswith("% Penetrasi"):
            nilai = current_section["% Penetrasi"]
            new_text = f"% Penetrasi: {format_percentage(nilai)}"

        elif text.startswith("Penjualan"):
            new_text = (
                f"Penjualan: "
                f"{format_desimal(current_section['Penjualan'])} kg"
            )

        elif text.startswith("Rata-rata Ticket Size"):
            new_text = (
                f"Rata-rata Ticket Size: "
                f"{format_desimal(current_section['Rata-rata Ticket Size'])} gram"
            )

        elif text.startswith("∆ MTD NoA") or text.startswith("Δ MTD NoA"):
            new_text = (
                f"∆ MTD NoA: "
                f"({format_ribuan(current_section['Δ MTD NoA'])}) NoA"
            )

        elif text.startswith("∆ MTD Penjualan") or text.startswith("Δ MTD Penjualan"):
            new_text = (
                f"∆ MTD Penjualan: "
                f"({format_desimal(current_section['Δ MTD Penjualan'])}) kg"
            )

        elif text.startswith("∆ MTD Ticket Size") or text.startswith("Δ MTD Ticket Size"):
            new_text = (
                f"∆ MTD Ticket Size: "
                f"({format_desimal(current_section['Δ MTD Ticket Size'])}) gr"
            )

        # =========================
        # UPDATE PARAGRAF
        # =========================
        if new_text:
            set_paragraph_text(para, new_text)
            jumlah_diubah += 1
            log.append(f"{text[:35]!r} -> {new_text}")

    return doc, jumlah_diubah, log


import streamlit as st

# =========================
# BSI THEME - CUSTOM CSS
# =========================
st.markdown("""
    <style>
    :root {
        --bsi-teal: #00A19C;
        --bsi-teal-dark: #006A67;
        --bsi-teal-darker: #004E4B;
        --bsi-gold: #F5A623;
        --bsi-gold-light: #FFD98A;
        --bsi-bg: #EAF6F5;
    }

    /* =========================================
       BACKGROUND UTAMA
       ========================================= */
    .stApp {
        background: linear-gradient(180deg, var(--bsi-bg) 0%, #FFFFFF 40%) !important;
    }

    /* Hanya container pembungkus teks yang ditransparankan,
       supaya tidak ada kotak putih menutupi tulisan,
       tapi widget (input, uploader, dataframe) tetap punya background sendiri */
    div[data-testid="stMarkdownContainer"],
    div[data-testid="stVerticalBlock"],
    div[data-testid="stHorizontalBlock"],
    div.element-container {
        background-color: transparent !important;
    }

    /* =========================================
       HEADER BANNER ALA LOGO BSI
       ========================================= */
    .bsi-header {
        background: linear-gradient(135deg, var(--bsi-teal-dark) 0%, var(--bsi-teal) 100%);
        padding: 28px 32px;
        border-radius: 14px;
        margin-bottom: 28px;
        box-shadow: 0 4px 14px rgba(0,106,103,0.35);
        display: flex;
        align-items: center;
        gap: 16px;
        border: 2px solid var(--bsi-gold);
    }
    .bsi-header .bsi-star {
        font-size: 34px;
        color: var(--bsi-gold);
        text-shadow: 0 0 12px rgba(245,166,35,0.6);
    }
    .bsi-header h1 {
        color: #FFFFFF !important;
        font-weight: 800 !important;
        margin: 0 !important;
        font-size: 26px !important;
    }
    .bsi-header p {
        color: var(--bsi-gold-light) !important;
        margin: 4px 0 0 0 !important;
        font-size: 14px !important;
    }

    /* =========================================
       HEADING NON-BANNER
       ========================================= */
    h1:not(.bsi-header h1) {
        color: var(--bsi-teal-darker) !important;
    }
    h2, h3 {
        color: var(--bsi-teal-dark) !important;
        border-left: 5px solid var(--bsi-gold);
        padding-left: 10px;
        margin-top: 18px !important;
    }

    /* =========================================
       TEKS st.markdown / st.write
       ========================================= */
    div[data-testid="stMarkdownContainer"] p,
    div[data-testid="stMarkdownContainer"] li,
    div[data-testid="stMarkdownContainer"] span {
        color: var(--bsi-teal-dark) !important;
    }
    div[data-testid="stMarkdownContainer"] strong {
        color: var(--bsi-teal-darker) !important;
    }

    /* Label semua widget (text_input, number_input, selectbox, dll) */
    label, [data-testid="stWidgetLabel"] p {
        color: var(--bsi-teal-darker) !important;
        font-weight: 600 !important;
    }

    /* Caption */
    [data-testid="stCaptionContainer"] p {
        color: #4A6B69 !important;
    }

    /* =========================================
       TOMBOL
       ========================================= */
    div.stButton > button[kind="primary"],
    div.stDownloadButton > button {
        background: linear-gradient(135deg, var(--bsi-teal) 0%, var(--bsi-teal-dark) 100%) !important;
        color: #FFFFFF !important;
        border: 2.5px solid var(--bsi-gold) !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        padding: 10px 22px !important;
        box-shadow: 0 3px 10px rgba(0,106,103,0.3);
        transition: 0.2s ease-in-out;
    }
    div.stButton > button[kind="primary"]:hover,
    div.stDownloadButton > button:hover {
        background: var(--bsi-gold) !important;
        color: #1a1a1a !important;
        border-color: var(--bsi-teal-dark) !important;
        transform: translateY(-1px);
    }
    div.stButton > button:not([kind="primary"]) {
        border: 2px solid var(--bsi-teal) !important;
        color: var(--bsi-teal-dark) !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        background-color: #FFFFFF !important;
    }

    /* =========================================
       FILE UPLOADER
       ========================================= */
    section[data-testid="stFileUploader"] {
        border: 3px dashed var(--bsi-gold) !important;
        border-radius: 12px;
        background-color: #FFFFFF !important;
        padding: 14px;
    }
    section[data-testid="stFileUploader"]:hover {
        border-color: var(--bsi-teal) !important;
        background-color: #F2FBFA !important;
    }

    /* =========================================
       EXPANDER
       ========================================= */
    .streamlit-expanderHeader {
        background: linear-gradient(90deg, var(--bsi-teal) 0%, var(--bsi-teal-dark) 100%) !important;
        color: #FFFFFF !important;
        border-radius: 8px;
        font-weight: 700;
        padding: 8px 12px !important;
        border: 1.5px solid var(--bsi-gold) !important;
    }
    .streamlit-expanderContent {
        background-color: #FFFFFF !important;
        border: 1px solid var(--bsi-teal) !important;
        border-top: none !important;
        border-radius: 0 0 8px 8px;
    }

    /* =========================================
       DIVIDER
       ========================================= */
    hr {
        border: none !important;
        height: 4px !important;
        background: linear-gradient(90deg, var(--bsi-gold), var(--bsi-teal)) !important;
        border-radius: 2px;
        margin: 20px 0 !important;
    }

    /* =========================================
       ALERT / INFO / SUCCESS BOX
       ========================================= */
    div[data-testid="stAlert"] {
        border-left: 6px solid var(--bsi-gold) !important;
        background-color: #E3F5F3 !important;
        border-radius: 6px;
    }
    div[data-testid="stAlert"] p {
        color: var(--bsi-teal-darker) !important;
    }
    div[data-baseweb="notification"] {
        background-color: #FFF7E8 !important;
        border-left: 6px solid var(--bsi-gold) !important;
    }

    /* =========================================
       SIDEBAR (jika ada)
       ========================================= */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, var(--bsi-teal-darker) 0%, var(--bsi-teal-dark) 100%) !important;
    }
    section[data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }

    /* =========================================
       DATAFRAME
       ========================================= */
    .stDataFrame {
        border: 1.5px solid var(--bsi-teal) !important;
        border-radius: 8px;
        overflow: hidden;
    }
    .stDataFrame thead tr th {
        background: linear-gradient(90deg, var(--bsi-teal-dark), var(--bsi-teal)) !important;
        color: #FFFFFF !important;
        font-weight: 700 !important;
    }
    .stDataFrame tbody tr:nth-child(even) {
        background-color: #F2FBFA !important;
    }
    .stDataFrame tbody tr:hover {
        background-color: var(--bsi-gold-light) !important;
    }

/* =========================================
       DATE INPUT
       ========================================= */
    div[data-baseweb="input"]:has(input[aria-label*="tanggal" i]),
    div[data-testid="stDateInput"] > div {
        background-color: #FFFFFF !important;
        border: 2px solid var(--bsi-gold) !important;
        border-radius: 6px !important;
    }
    div[data-testid="stDateInput"] input {
        color: var(--bsi-teal-darker) !important;
        font-weight: 500 !important;
    }
    div[data-testid="stDateInput"] svg {
        fill: var(--bsi-teal) !important;
    }

    /* Angka stepper number_input */
    button[data-testid="stNumberInputStepUp"],
    button[data-testid="stNumberInputStepDown"] {
        background-color: var(--bsi-teal) !important;
        color: #FFFFFF !important;
    }

    /* =========================================
       COLUMN SPACING RAPI
       ========================================= */
    div[data-testid="stHorizontalBlock"] {
        gap: 1.2rem;
    }
    /* =========================================
       FIX: TEKS DI DALAM TOMBOL & BANNER KETIMPA
       WARNA TEAL DARI RULE MARKDOWN GENERIK
       (Streamlit membungkus teks tombol/banner
       dalam stMarkdownContainer yang sama, jadi
       perlu selector lebih spesifik utk menang)
       ========================================= */

    /* Judul & subjudul banner header - paksa putih & emas */
    div[data-testid="stMarkdownContainer"] .bsi-header h1,
    .bsi-header div[data-testid="stMarkdownContainer"] h1 {
        color: #FFFFFF !important;
        -webkit-text-fill-color: #FFFFFF !important;
    }
    div[data-testid="stMarkdownContainer"] .bsi-header p,
    .bsi-header div[data-testid="stMarkdownContainer"] p {
        color: var(--bsi-gold-light) !important;
        -webkit-text-fill-color: var(--bsi-gold-light) !important;
    }

    /* Teks di dalam tombol primary (Generate Dokumen Word) - paksa putih */
    div.stButton > button[kind="primary"] p,
    div.stButton > button[kind="primary"] div[data-testid="stMarkdownContainer"] p,
    div.stButton > button[kind="primary"] span {
        color: #FFFFFF !important;
        -webkit-text-fill-color: #FFFFFF !important;
        font-weight: 700 !important;
    }
    div.stButton > button[kind="primary"]:hover p,
    div.stButton > button[kind="primary"]:hover div[data-testid="stMarkdownContainer"] p {
        color: #1a1a1a !important;
        -webkit-text-fill-color: #1a1a1a !important;
    }

    /* Teks di dalam tombol download (Download Hasil Word) - paksa putih */
    div.stDownloadButton > button p,
    div.stDownloadButton > button div[data-testid="stMarkdownContainer"] p,
    div.stDownloadButton > button span {
        color: #FFFFFF !important;
        -webkit-text-fill-color: #FFFFFF !important;
        font-weight: 700 !important;
    }
    div.stDownloadButton > button:hover p,
    div.stDownloadButton > button:hover div[data-testid="stMarkdownContainer"] p {
        color: #1a1a1a !important;
        -webkit-text-fill-color: #1a1a1a !important;
    }

    /* Teks di dalam tombol secondary (non-primary) - paksa teal-dark */
    div.stButton > button:not([kind="primary"]) p,
    div.stButton > button:not([kind="primary"]) div[data-testid="stMarkdownContainer"] p {
        color: var(--bsi-teal-dark) !important;
        -webkit-text-fill-color: var(--bsi-teal-dark) !important;
    }
    </style>
""", unsafe_allow_html=True)

# =========================
# HEADER BANNER (ala logo BSI)
# =========================
st.markdown("""
    <div class="bsi-header">
        <div class="bsi-star">✦</div>
        <div>
            <h1>Auto Wording Segmen Excel to Word</h1>
            <p>Upload file Excel & template Word setiap kali laporan baru terbit</p>
        </div>
    </div>
""", unsafe_allow_html=True)

# =========================
# UI
# =========================
col_upload1, col_upload2 = st.columns(2)
with col_upload1:
    excel_file = st.file_uploader("1️⃣ Upload file Excel (.xlsx)", type=["xlsx"])
with col_upload2:
    word_file = st.file_uploader("2️⃣ Upload template Word (.docx)", type=["docx"])

if excel_file:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet_names = wb.sheetnames
    default_idx = sheet_names.index("REGION") if "REGION" in sheet_names else 0
    sheet_name = st.selectbox("Pilih sheet", sheet_names, index=default_idx)
    sheet = wb[sheet_name]

        # ===== TANGGAL LAPORAN =====
    st.subheader("📅 Tanggal Laporan")
    tanggal_laporan = st.date_input(
        "Pilih tanggal laporan",
        value=None,
        format="DD/MM/YYYY",
        help="Tanggal ini akan otomatis menggantikan tanggal 'sd ...' di template Word.",
    )
    if tanggal_laporan:
        st.info(
            f"Tanggal yang akan digunakan di Word: "
            f"**{format_tanggal_indonesia(tanggal_laporan)}**"
        )

    st.divider()

    st.subheader("🔎 Deteksi Baris per Segmen")
    st.markdown("**Kata kunci pencarian label per segmen** (ubah jika istilah di Excel berbeda)")
    kw_cols = st.columns(4)
    keyword_mapping = {}
    for i, segmen in enumerate(SEGMEN_LIST):
        with kw_cols[i]:
            keyword_mapping[segmen] = st.text_input(
                f"Kata kunci - {segmen}", value=segmen, key=f"kw_{segmen}"
            )

    grand_total_keyword = st.text_input(
        "Kata kunci baris total data (baris yang diambil sebagai DATA segmen)",
        value="Grand Total",
        help="Ubah jika di Excel Anda istilahnya berbeda, misal 'Total' saja.",
    )

    cache_key = (excel_file.name, sheet_name, tuple(keyword_mapping.values()), grand_total_keyword)
    if st.session_state.get("_cache_key") != cache_key:
        label_rows = cari_baris_segmen(sheet, keyword_mapping)
        st.session_state.label_rows = label_rows
        st.session_state.auto_rows = cari_baris_grand_total(
            sheet, label_rows, grand_total_keyword
        )
        st.session_state._cache_key = cache_key

    label_rows = st.session_state.label_rows
    auto_rows = st.session_state.auto_rows

    st.markdown("**Baris data (baris 'Grand Total' yang ditemukan, bisa dikoreksi manual)**")
    baris_input = {}
    cols = st.columns(4)
    for i, segmen in enumerate(SEGMEN_LIST):
        label_row = label_rows.get(segmen)
        data_row = auto_rows.get(segmen)
        if label_row is None:
            info = "⚠️ label tidak ditemukan"
        elif data_row is None:
            info = f"label di baris {label_row}, tapi '{grand_total_keyword}' tidak ditemukan"
        else:
            info = f"label di baris {label_row} → Grand Total di baris {data_row}"
        with cols[i]:
            st.markdown(f"**{segmen}**")
            st.caption(info)
            default_val = data_row if data_row else (label_row if label_row else 1)
            baris_input[segmen] = st.number_input(
                f"Baris data {segmen}",
                min_value=1,
                value=default_val,
                key=f"row_{segmen}_{cache_key}",
            )

    with st.expander("⚙️ Pengaturan lanjutan: mapping kolom"):
        st.caption("Ubah huruf kolom jika struktur Excel berbeda dari biasanya.")
        kolom_mapping = {}
        for label, default_kolom in DEFAULT_KOLOM.items():
            kolom_mapping[label] = st.text_input(label, value=default_kolom, key=f"col_{label}")

    data_per_segmen = {
        segmen: get_row_data(sheet, baris_input[segmen], kolom_mapping)
        for segmen in SEGMEN_LIST
    }

    st.subheader("📋 Preview Data")
    df_preview = pd.DataFrame(data_per_segmen).T
    st.dataframe(df_preview, use_container_width=True)

    st.divider()

    if word_file:
        if st.button("Generate Dokumen Word", type="primary"):
            doc = Document(word_file)
            doc, jumlah_diubah, log = proses_dokumen(
            doc,
            data_per_segmen,
            tanggal_laporan=tanggal_laporan
            )

            buffer = io.BytesIO()
            doc.save(buffer)
            buffer.seek(0)

            st.success(f"Selesai! {jumlah_diubah} paragraf berhasil diubah.")
            with st.expander("Lihat detail perubahan"):
                for line in log:
                    st.text(line)

            nama_output = re.sub(r"\.xlsx$", "", excel_file.name, flags=re.IGNORECASE)
            st.download_button(
                "⬇️ Download Hasil Word",
                data=buffer,
                file_name=f"Revised_{nama_output}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
    else:
        st.info("Upload template Word untuk melanjutkan.")
else:
    st.info("Silakan upload file Excel terlebih dahulu.")